"""
_advisor.py — Parse the Student Advisor Listing spreadsheet and provide lookup.

Parses the semi-structured advisor assignment spreadsheet into a flat list of
records, each mapping a (program, campus, last-name-range) to an advisor with
contact details.  Also loads program codes from Programs.xlsx for code/name
matching and autocomplete.
"""

import logging
import re
import urllib.request
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
ADVISOR_FILE = REPO_ROOT / "assets" / "spreadsheets" / "Student_Advisor_Listing_current_1.xlsx"
PROGRAMS_FILE = REPO_ROOT / "assets" / "spreadsheets" / "Programs.xlsx"

# Map campus shorthand codes to canonical names found in the spreadsheet.
CAMPUS_CODES = {
    "LMC": ["Lee", "Sanford", "Lee "],
    "HMC": ["Harnett", "Lillington"],
    "PMC": ["Pittsboro", "Chatham", "Chatham Health"],
    "CMC": ["Chatham Main", "Chatham Campus"],
    "DUNN": ["Dunn"],
    "ESTC": ["ESTC"],
    "WHC": ["West Harnett Center"],
    "REMOTE": ["Remote", "Virtual"],
}

# Reverse map: canonical campus name (lowered) -> short code
_CAMPUS_REVERSE: dict[str, str] = {}
for code, names in CAMPUS_CODES.items():
    for name in names:
        _CAMPUS_REVERSE[name.strip().lower()] = code


def campus_code_for(campus_name: str) -> str:
    """Return the short campus code (uppercase) for a campus name, or the name itself."""
    return _CAMPUS_REVERSE.get(campus_name.strip().lower(), campus_name)


def _campus_code_from_text(text: str) -> str:
    """Try to extract a campus code from free-form text.

    Handles messy values like 'Located on PMC, also serves LMC', '122.0',
    and compound headers like 'Lee - GOT Nursing'.
    """
    t = text.strip().lower()
    # Check exact match first
    code = _CAMPUS_REVERSE.get(t)
    if code:
        return code
    # Check for short campus codes in the text (e.g., "PMC", "LMC", "HMC")
    for short in ("LMC", "HMC", "PMC", "CMC", "ESTC", "WHC"):
        if short.lower() in t:
            return short
    # Check for canonical campus name substrings (longest first to prefer specific)
    for name in sorted(_CAMPUS_REVERSE.keys(), key=len, reverse=True):
        if name in t:
            return _CAMPUS_REVERSE[name]
    return ""


# ---------------------------------------------------------------------------
# Name-range parsing — multi-character prefix ranges
# ---------------------------------------------------------------------------

def _parse_name_ranges(text: str) -> list[tuple[str, str]]:
    """Extract list of (start_prefix, end_prefix) from a name-range string.

    Returns uppercase prefix pairs for precision matching.  Handles:
      "Last Names A-B"        → [("A", "B")]
      "Last Names Gr-Gu"      → [("GR", "GU")]
      "Last Names G-K, N-R"   → [("G", "K"), ("N", "R")]
      "All Students"           → [("A", "Z")]
      "Preparatory Dental (A-K) A55280DH" → [("A", "K")]
    """
    t = text.strip()

    # "All Students" / "1st Year" etc — matches everything
    if re.match(
        r"(?i)(all\s+students?|all\s+campuses?|all\s+admitted|"
        r"1st\s+year|2nd\s+year|full.time|part.time)",
        t,
    ):
        return [("A", "Z")]

    # Find all letter-range patterns: "A-B", "Gr-Gu", "A - Z"
    # Only accept short prefixes (1-2 chars) to avoid false matches on
    # hyphenated words like "Arts-Teacher", "Pre-Medical", "Arts-Pittsboro".
    ranges = re.findall(r"\b([A-Z][a-z]?)\s*[-–—]\s*([A-Z][a-z]?)\b", t)
    if ranges:
        return [(s.upper(), e.upper()) for s, e in ranges]

    # Single letter: "Last Names E", "Last Names C"
    m = re.search(r"(?:last\s*names?\s*(?:of\s*)?)\s*([A-Z])\s*$", t, re.IGNORECASE)
    if m:
        letter = m.group(1).upper()
        return [(letter, letter)]

    # Fallback — treat as all
    return [("A", "Z")]


def _name_in_ranges(last_name: str, ranges: list[tuple[str, str]]) -> bool:
    """Check if a last name falls within any of the given prefix ranges.

    Uses string comparison for multi-character precision:
      "GRANT" in ("GR", "GU") → True
      "GATES" in ("GR", "GU") → False
      "GUTHRIE" in ("GR", "GU") → True (starts with "GU")
    """
    n = last_name.strip().upper()
    if not n:
        return False
    for start, end in ranges:
        if n >= start and (n <= end or n.startswith(end)):
            return True
    return False


# ---------------------------------------------------------------------------
# Embedded program code extraction
# ---------------------------------------------------------------------------

_CODE_RE = re.compile(r"\b([ADCT]\d{5}[A-Z]{0,4})\b")


def _extract_codes(text: str) -> list[str]:
    """Extract program codes like A55280NR, D45240, C55120 from text."""
    return _CODE_RE.findall(text.upper())


# ---------------------------------------------------------------------------
# Advisor name cleanup
# ---------------------------------------------------------------------------

def _clean_advisor_name(name: str) -> str:
    """Strip informal notes from advisor name text."""
    if not name:
        return ""
    # Known corrections for first-name-only entries
    _NAME_CORRECTIONS = {
        "brenda": "Brenda Grubb",
        "lisa": "Lisa Smelser",
    }
    corrected = _NAME_CORRECTIONS.get(name.strip().lower())
    if corrected:
        return corrected
    # Not really a person's name
    if re.search(r"(?i)once officially admitted|will be assigned", name):
        return ""
    # Remove trailing parenthetical notes like "(only)"
    cleaned = re.sub(r"\s*\(only\)\s*$", "", name, flags=re.IGNORECASE)
    # Remove trailing action notes like "will assign"
    cleaned = re.sub(r"\s+will\s+assign\s*$", "", cleaned, flags=re.IGNORECASE)
    return cleaned.strip()


# ---------------------------------------------------------------------------
# Row classification
# ---------------------------------------------------------------------------

def _is_program_header(row_vals: list) -> bool:
    """Detect a program header row: column A has text, B-F are None."""
    a = row_vals[0]
    if not a or not isinstance(a, str):
        return False
    a = a.strip()
    if not a:
        return False
    if any(row_vals[i] is not None for i in range(1, min(len(row_vals), 6))):
        return False
    if "advisor assignments by program" in a.lower():
        return False
    return True


def _is_campus_header(row_vals: list) -> bool:
    """Detect a campus/sub-header row: col B == 'Advisor'."""
    b = row_vals[1] if len(row_vals) > 1 else None
    return isinstance(b, str) and b.strip().lower() == "advisor"


def _is_data_row(row_vals: list) -> bool:
    """Detect a data row: col A has last-name-range text, col B has an advisor name."""
    a = row_vals[0]
    b = row_vals[1] if len(row_vals) > 1 else None
    if not a or not isinstance(a, str):
        return False
    if not b or not isinstance(b, str):
        return bool(re.search(r"(?i)(last\s*name|all\s*student|1st\s*year|2nd\s*year)", a))
    return True


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------

def parse_advisor_spreadsheet(filepath: Path | None = None) -> list[dict]:
    """Parse the advisor spreadsheet into a flat list of advisor records.

    Each record:
        {
            "program": str,
            "campus": str,
            "campus_code": str,
            "name_range_text": str,
            "name_ranges": list[tuple[str, str]],  # prefix ranges
            "range_start": str,   # outer-bound letter (compat)
            "range_end": str,     # outer-bound letter (compat)
            "advisor_name": str,
            "advisor_id": str,
            "office": str,
            "email": str,
            "embedded_codes": list[str],  # program codes in data row
        }
    """
    filepath = filepath or ADVISOR_FILE
    if not filepath.exists():
        return []

    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    ws = wb.active
    records = []
    current_program = ""
    current_campus = ""

    for row in ws.iter_rows(max_col=6, values_only=True):
        vals = list(row) + [None] * (6 - len(row))  # pad to 6 cols

        if _is_program_header(vals):
            current_program = vals[0].strip()
            current_campus = ""
            continue

        if _is_campus_header(vals):
            campus_text = vals[0]
            if campus_text and isinstance(campus_text, str) and campus_text.strip():
                raw = campus_text.strip()
                # Try to extract a clean campus name from compound headers
                code = _campus_code_from_text(raw)
                if code:
                    # Use the first canonical name for this code
                    current_campus = CAMPUS_CODES[code][0]
                else:
                    current_campus = raw
            continue

        if not current_program:
            continue

        # Data row
        a = vals[0]
        if not a or not isinstance(a, str):
            continue
        a = a.strip()
        if not a:
            continue

        raw_name = vals[1].strip() if isinstance(vals[1], str) else ""
        advisor_name = _clean_advisor_name(raw_name)
        advisor_id = str(int(vals[2])) if vals[2] is not None else ""
        office = vals[3].strip() if isinstance(vals[3], str) else ""
        email = vals[5].strip() if isinstance(vals[5], str) else ""

        # Campus: prefer col E, fall back to current_campus from header
        campus_raw = vals[4].strip() if isinstance(vals[4], str) else ""
        # Skip numeric garbage like "122.0"
        if campus_raw and not re.match(r"^\d", campus_raw):
            campus = campus_raw
        elif current_campus:
            campus = current_campus
        else:
            campus = ""

        # Try to improve campus matching for messy col E values
        campus_code = _campus_code_from_text(campus) if campus else ""
        if not campus_code and campus:
            campus_code = campus_code_for(campus)

        # Skip header-like rows that slipped through
        if a.lower() in ("advisor", "email", "office location", "campus"):
            continue

        # Skip if this looks like a campus header variant
        if not advisor_name and not advisor_id:
            if not re.search(r"(?i)(last\s*name|all\s*student|1st|2nd|preparatory|pre\s)", a):
                continue

        # Parse name ranges with multi-char prefix precision
        name_ranges = _parse_name_ranges(a)
        # Outer-bound single letters for backward compat
        all_starts = [r[0][0] for r in name_ranges]
        all_ends = [r[1][0] for r in name_ranges]
        range_start = min(all_starts) if all_starts else "A"
        range_end = max(all_ends) if all_ends else "Z"

        # Extract embedded program codes from the data row
        embedded_codes = _extract_codes(a)

        # Split "X or Y" advisor entries into separate records
        _SPLIT_ADVISORS = {
            "brenda or lisa": [
                {"name": "Brenda Grubb"},
                {"name": "Lisa Smelser"},
            ],
        }
        split_key = raw_name.strip().lower()
        advisor_list = _SPLIT_ADVISORS.get(split_key)
        if advisor_list:
            for entry in advisor_list:
                records.append({
                    "program": current_program,
                    "campus": campus,
                    "campus_code": campus_code,
                    "name_range_text": a,
                    "name_ranges": name_ranges,
                    "range_start": range_start,
                    "range_end": range_end,
                    "advisor_name": entry["name"],
                    "advisor_id": advisor_id,
                    "office": office,
                    "email": email,
                    "embedded_codes": embedded_codes,
                })
        else:
            records.append({
                "program": current_program,
                "campus": campus,
                "campus_code": campus_code,
                "name_range_text": a,
                "name_ranges": name_ranges,
                "range_start": range_start,
                "range_end": range_end,
                "advisor_name": advisor_name,
                "advisor_id": advisor_id,
                "office": office,
                "email": email,
                "embedded_codes": embedded_codes,
            })

    wb.close()
    return records


def _program_matches(query: str, record: dict) -> bool:
    """Check if a program query (code or name fragment) matches a record.

    Checks both the program header text and any embedded codes in the data row.
    """
    q = query.strip().lower()
    advisor_program = record["program"]
    ap = advisor_program.lower()

    # Extract codes from the advisor program header (inside parens, split by /)
    codes_in_parens = re.findall(r"\(([^)]+)\)", advisor_program)
    all_codes = []
    for group in codes_in_parens:
        all_codes.extend(c.strip().lower() for c in group.split("/"))

    # Also include embedded codes from the data row itself
    all_codes.extend(c.lower() for c in record.get("embedded_codes", []))

    # Check if query is a known program code
    programs = get_programs()
    query_is_code = any(p["code"].lower() == q for p in programs)

    if query_is_code:
        # Strict: the exact code must appear in combined code list
        return q in all_codes

    # Query is a name/text fragment — check two things:
    # 1) Direct substring match on the program name text
    if q in ap:
        return True

    # 2) Resolve matching program codes and check if they appear
    for p in programs:
        if q in p["name"].lower() or q in p["credential"].lower():
            if p["code"].lower() in all_codes:
                return True

    return False


def _filter(
    records: list[dict],
    last_name: str,
    campus_code: str = "",
    program: str = "",
) -> list[dict]:
    """Core filter — return records matching last name, optional campus & program.

    Uses multi-character prefix range matching for precision.
    """
    campus_names = []
    if campus_code:
        campus_names = [n.lower() for n in CAMPUS_CODES.get(campus_code.upper(), [campus_code])]

    matches = []
    for rec in records:
        if not _name_in_ranges(last_name, rec.get("name_ranges", [])):
            continue
        if campus_names:
            rec_campus = rec["campus"].strip().lower()
            if not any(cn in rec_campus or rec_campus in cn for cn in campus_names):
                continue
        if program and not _program_matches(program, rec):
            continue
        if not rec["advisor_name"]:
            continue
        matches.append(rec)
    return matches


def lookup_advisor(
    records: list[dict],
    last_name: str,
    campus_code: str = "",
    program: str = "",
) -> list[dict]:
    """Find matching advisor records for a student last name and optional campus/program.

    Always returns results — progressively broadens the search if needed:
    1. Exact filters (campus + program)
    2. Drop campus, keep program (program match is higher priority)
    3. Try campus = lmc with program
    4. Drop program, keep campus
    5. Try campus = lmc without program
    6. Drop all filters (name match only)

    Returns a list of matching records sorted by program.
    """
    if not last_name:
        return []

    # Build a cascade of progressively looser filter combos.
    # Program is prioritised over campus when broadening.
    attempts = [(campus_code, program)]
    if program:
        # Drop campus but keep program
        if campus_code:
            attempts.append(("", program))
        # Try LMC default with program
        if campus_code.upper() != "LMC":
            attempts.append(("LMC", program))
    if campus_code:
        # Drop program, keep original campus
        attempts.append((campus_code, ""))
    if campus_code.upper() != "LMC":
        # Try LMC default without program
        attempts.append(("LMC", ""))
    attempts.append(("", ""))  # everything

    for c, p in attempts:
        matches = _filter(records, last_name, c, p)
        if matches:
            matches.sort(key=lambda r: r["program"].lower())
            return matches

    # Ultimate fallback — ignore name ranges entirely.
    # Priority: program match first, then LMC campus.
    if program:
        matches = [
            r for r in records
            if r["advisor_name"] and _program_matches(program, r)
        ]
        if matches:
            matches.sort(key=lambda r: r["program"].lower())
            return matches
    # Default to LMC advisors
    matches = [
        r for r in records
        if r["advisor_name"] and r.get("campus_code") == "LMC"
    ]
    if matches:
        matches.sort(key=lambda r: r["program"].lower())
        return matches

    return []


# ---------------------------------------------------------------------------
# Contact enrichment — fill gaps from cccc.edu directory
# ---------------------------------------------------------------------------

# Known emails scraped from cccc.edu for advisors missing contact info
_DIRECTORY_EMAILS: dict[str, str] = {
    "Billy Freeman": "bfreeman@cccc.edu",
    "Steve Heesacker": "dhees901@cccc.edu",
    "John Wilson": "jwils563@cccc.edu",
    "Tiffany Needham": "tneed920@cccc.edu",
    "Roy Allen": "rallen@cccc.edu",
    "Lisa Smelser": "lsmel384@cccc.edu",
    "Brenda Grubb": "bgrub218@cccc.edu",
}

log = logging.getLogger(__name__)

# Override campus for advisors whose spreadsheet data is wrong/missing
_CAMPUS_OVERRIDES: dict[str, str] = {
    "Roy Allen": "ESTC",
    "Kris Rixon": "CMC",
}


def _enrich_records(records: list[dict]) -> None:
    """Fill in missing emails and fix campus assignments from fallback tables."""
    for rec in records:
        name = rec["advisor_name"]
        if not rec.get("email") and name in _DIRECTORY_EMAILS:
            rec["email"] = _DIRECTORY_EMAILS[name]
            log.debug("Enriched email %s → %s", name, rec["email"])
        if name in _CAMPUS_OVERRIDES:
            code = _CAMPUS_OVERRIDES[name]
            rec["campus"] = CAMPUS_CODES[code][0]
            rec["campus_code"] = code


# Cached records — loaded once per process
_CACHED_RECORDS: list[dict] | None = None


def get_records() -> list[dict]:
    """Return cached advisor records, parsing on first call."""
    global _CACHED_RECORDS
    if _CACHED_RECORDS is None:
        _CACHED_RECORDS = parse_advisor_spreadsheet()
        _enrich_records(_CACHED_RECORDS)
    return _CACHED_RECORDS


def reload_records() -> list[dict]:
    """Force re-parse of the spreadsheet (e.g. after upload of new version)."""
    global _CACHED_RECORDS
    _CACHED_RECORDS = parse_advisor_spreadsheet()
    return _CACHED_RECORDS


# ---------------------------------------------------------------------------
# Program codes from Programs.xlsx
# ---------------------------------------------------------------------------

# Words/acronyms that should stay uppercase in program names
_UPPER_WORDS = {
    "AAS", "AATP", "ASTP", "ACHR", "ADN", "BLET", "CAT", "CCP", "FAST",
    "GOT", "HRM", "IC3", "IST", "PTA", "RN", "PN", "VMT", "EV",
}

# Minor words kept lowercase (unless first word)
_MINOR_WORDS = {"a", "an", "and", "at", "by", "for", "from", "in", "of", "on", "or", "the", "to"}


def _clean_program_name(name: str) -> str:
    """Normalize a program name for consistent display.

    - Title-case words (preserving known acronyms, lowering minor words)
    - Replace hyphens used as word separators with spaces
    - Strip embedded program codes (e.g. '- C20100K1')
    - Fix unbalanced parentheses
    - Replace underscores with spaces
    - Collapse whitespace
    - Fix known typos
    """
    # Strip trailing program code references like '- C20100K1' or 'C20100K1' (case-insensitive)
    name = re.sub(r"\s*-?\s*[ADCT]\d{5}[A-Z\d]{0,4}\s*$", "", name, flags=re.IGNORECASE)
    # Replace underscores with spaces
    name = name.replace("_", " ")
    # Fix known typos
    name = re.sub(r"(?i)resipiratory", "Respiratory", name)
    # Normalize " -CCP" → " CCP", "Tech-PTA" → "Tech PTA" (hyphen before acronym/word)
    name = re.sub(r"\s*-\s*(?=[A-Za-z])", " ", name)
    # Fix "ADV." → "Advanced"
    name = re.sub(r"\bADV\.\s*", "Advanced ", name, flags=re.IGNORECASE)
    # Fix "Tech." → "Tech"
    name = re.sub(r"\bTech\.", "Tech", name, flags=re.IGNORECASE)
    # Collapse whitespace
    name = re.sub(r"\s+", " ", name).strip()
    # Fix unbalanced parentheses
    opens = name.count("(")
    closes = name.count(")")
    if opens > closes:
        name += ")" * (opens - closes)
    elif closes > opens:
        # Remove excess closing parens from the end
        while name.count(")") > name.count("(") and name.endswith(")"):
            name = name[:-1]
        name = name.strip()
    # Title case: capitalize each word, preserve acronyms, lower minor words
    words = name.split()
    result = []
    for i, w in enumerate(words):
        # Strip non-alphanumeric for acronym check
        stripped = re.sub(r"[^A-Za-z0-9]", "", w).upper()
        # Check if the core is a known acronym
        if stripped in _UPPER_WORDS:
            result.append(re.sub(r"[A-Za-z]+", lambda m: m.group().upper(), w))
        elif i > 0 and w.lower() in _MINOR_WORDS:
            result.append(w.lower())
        elif w.isupper() and len(w) > 2:
            # ALL-CAPS non-acronym → capitalize
            result.append(w.capitalize())
        elif w[0].islower():
            # Lowercase start → capitalize
            result.append(w.capitalize())
        else:
            result.append(w)
    name = " ".join(result)
    return name


# ---------------------------------------------------------------------------
# Photo URL map from synced faculty-staff directory
# ---------------------------------------------------------------------------

_PLACEHOLDER_PHOTO = "IMG_Roster_Placeholder.png"
_PHOTO_RE = re.compile(
    r'<img\s+src="([^"]+)"[^>]*>\s*\|\s*\[([^\]]+)\]',
    re.IGNORECASE,
)


def _parse_directory_entries() -> list[dict]:
    """Parse faculty-staff-directory.md → list of {name, photo_url, role, profile_url}.

    Each row in the directory tables has: photo | [Name](url) | Role
    """
    md_path = REPO_ROOT / "documentation" / "faculty-staff-directory.md"
    if not md_path.exists():
        return []
    text = md_path.read_text(encoding="utf-8", errors="replace")
    # Match: | <img src="photo"> | [Name](profile_url) | Role |
    row_re = re.compile(
        r'<img\s+src="([^"]+)"[^>]*>\s*\|\s*\[([^\]]+)\]\(([^)]+)\)\s*\|\s*([^|\n]+)',
        re.IGNORECASE,
    )
    entries: list[dict] = []
    for m in row_re.finditer(text):
        photo_url = m.group(1).strip()
        name = m.group(2).strip()
        profile_url = m.group(3).strip()
        role = m.group(4).strip()
        if _PLACEHOLDER_PHOTO in photo_url:
            photo_url = ""
        entries.append({
            "name": name,
            "photo_url": photo_url,
            "profile_url": profile_url,
            "role": role,
        })
    return entries


def get_photo_map() -> dict[str, str]:
    """Parse faculty-staff-directory.md → {normalised_name: photo_url}.

    Excludes entries that use the placeholder headshot.
    """
    entries = _parse_directory_entries()
    return {e["name"].lower(): e["photo_url"] for e in entries if e["photo_url"]}


def get_navigators() -> list[dict]:
    """Return Education Navigators from the faculty-staff directory.

    Matches anyone whose role contains 'education navigator' (case-insensitive).
    Returns list of {name, photo_url, profile_url, role}, sorted by name.
    """
    entries = _parse_directory_entries()
    navs = [e for e in entries if "education navigator" in e["role"].lower()]
    return sorted(navs, key=lambda e: e["name"].rsplit(None, 1)[-1].lower())


def get_advisor_directory() -> list[dict]:
    """Return a deduplicated, alphabetical list of unique advisors.

    Each entry: {name, email, office, campuses, programs, photo_url}.
    """
    records = get_records()
    advisors: dict[str, dict] = {}
    for rec in records:
        name = rec["advisor_name"]
        if not name:
            continue
        if name not in advisors:
            advisors[name] = {
                "name": name,
                "advisor_id": rec.get("advisor_id", ""),
                "email": rec.get("email", ""),
                "office": rec.get("office", ""),
                "campuses": set(),
                "programs": set(),
            }
        entry = advisors[name]
        # Keep the best data (prefer non-empty)
        if not entry["email"] and rec.get("email"):
            entry["email"] = rec["email"]
        if not entry["office"] and rec.get("office"):
            entry["office"] = rec["office"]
        if rec.get("campus"):
            entry["campuses"].add(rec["campus"])
        if rec.get("program"):
            entry["programs"].add(rec["program"])
    # Convert sets to sorted lists for JSON serialization
    photos = get_photo_map()
    result = []
    for adv in sorted(advisors.values(), key=lambda a: a["name"].rsplit(None, 1)[-1].lower()):
        adv["campuses"] = sorted(adv["campuses"])
        adv["programs"] = sorted(adv["programs"])
        adv["photo_url"] = photos.get(adv["name"].lower(), "")
        result.append(adv)
    return result


# Preferred campus group ordering and display names
_CAMPUS_ORDER = ["LMC", "HMC", "PMC", "CMC", "DUNN", "ESTC", "WHC", "REMOTE"]
_CAMPUS_LABELS = {
    "LMC": "Lee Main Campus (LMC)",
    "HMC": "Harnett Main Campus (HMC)",
    "PMC": "Pittsboro Main Campus (PMC)",
    "CMC": "Chatham Main Campus (CMC)",
    "DUNN": "Dunn Center",
    "ESTC": "ESTC",
    "WHC": "West Harnett Center (WHC)",
    "REMOTE": "Remote / Virtual",
}


def get_advisor_directory_grouped() -> list[dict]:
    """Return advisor directory grouped by primary campus.

    Returns a list of dicts: {code, label, advisors}.
    Groups are ordered by _CAMPUS_ORDER; advisors alphabetically within each.
    """
    directory = get_advisor_directory()
    groups: dict[str, list[dict]] = {}
    for adv in directory:
        if adv["campuses"]:
            code = campus_code_for(adv["campuses"][0])
            # Handle messy values that don't resolve to a known code
            if code not in _CAMPUS_LABELS:
                code = "OTHER"
        else:
            code = "OTHER"
        groups.setdefault(code, []).append(adv)

    # Build ordered result
    result = []
    for code in _CAMPUS_ORDER:
        if code in groups:
            result.append({
                "code": code,
                "label": _CAMPUS_LABELS[code],
                "advisors": groups.pop(code),
            })
    # Remaining (OTHER or unknown)
    for code, advisors in sorted(groups.items()):
        result.append({
            "code": code,
            "label": _CAMPUS_LABELS.get(code, "Other"),
            "advisors": advisors,
        })
    return result


_CACHED_PROGRAMS: list[dict] | None = None


def parse_programs(filepath: Path | None = None) -> list[dict]:
    """Parse the Programs.xlsx 'Program Codes' sheet.

    Returns a list of dicts with keys: code, name, credential, level.
    """
    filepath = filepath or PROGRAMS_FILE
    if not filepath.exists():
        return []

    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    ws = wb["Program Codes"]
    programs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row) + [None] * (10 - len(row))
        code = vals[0]
        if not code or not isinstance(code, str):
            continue
        code = code.strip()
        name = _clean_program_name(vals[1].strip()) if isinstance(vals[1], str) else code
        credential = vals[2].strip() if isinstance(vals[2], str) else ""
        level = vals[3].strip() if isinstance(vals[3], str) else ""
        programs.append({
            "code": code,
            "name": name,
            "credential": credential,
            "level": level,
        })
    wb.close()

    # Sort alphabetically by code, grouped by numeric core so related
    # programs (e.g., A25800, C25800, C25800PO, D25800) stay together.
    def _program_sort_key(p: dict) -> tuple[str, str]:
        m = re.match(r"[A-Z](\d{5})", p["code"])
        numeric = m.group(1) if m else ""
        return (numeric, p["code"])

    programs.sort(key=_program_sort_key)
    return programs


def get_programs() -> list[dict]:
    """Return cached program list, parsing on first call."""
    global _CACHED_PROGRAMS
    if _CACHED_PROGRAMS is None:
        _CACHED_PROGRAMS = parse_programs()
    return _CACHED_PROGRAMS
